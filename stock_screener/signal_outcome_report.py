from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from stock_screener.signal_outcome_study import SignalOutcomeStudyResult


HEADER_FILL = PatternFill("solid", fgColor="DFF5EE")
BOLD_FONT = Font(bold=True)


def write_signal_outcome_workbook(result: SignalOutcomeStudyResult, workbook_path: Path) -> Path:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    _write_summary_sheet(summary_sheet, result.summary)
    _write_dataframe_sheet(workbook, "Stock Outcome Stats", _format_dates(result.stock_stats))

    workbook.save(workbook_path)
    return workbook_path


def _write_summary_sheet(sheet: Any, summary: dict[str, Any]) -> None:
    rows = [
        ("Exchange", summary.get("exchange", "")),
        ("Signal scope", summary.get("signal_scope", "")),
        ("Target gain %", summary.get("target_gain_pct", 0)),
        ("Current signal universe", summary.get("current_signal_universe_count", 0)),
        ("Current BUY", summary.get("current_buy_count", 0)),
        ("Current SELL", summary.get("current_sell_count", 0)),
        ("Historical pairs analyzed", summary.get("historical_pairs_analyzed", 0)),
        ("Avg target hit rate %", summary.get("avg_target_hit_rate_pct", 0)),
        ("Median days to target", summary.get("median_days_to_target", 0)),
        ("Median peak gain %", summary.get("median_peak_gain_pct", 0)),
        ("Avg failed BUY rate %", summary.get("avg_failed_buy_rate_pct", 0)),
    ]
    for row_index, (label, value) in enumerate(rows, start=1):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=_excel_value(value))
        sheet.cell(row=row_index, column=1).font = BOLD_FONT
    _finish_sheet(sheet, widths={"A": 28, "B": 20})


def _write_dataframe_sheet(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(title)
    if frame.empty:
        sheet.append(["No rows"])
        _finish_sheet(sheet)
        return

    sheet.append(list(frame.columns))
    for _, row in frame.iterrows():
        sheet.append([_excel_value(value) for value in row.tolist()])

    for cell in sheet[1]:
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    date_columns = {"latest_week_date", "current_signal_date"}
    for column_index, cell in enumerate(sheet[1], start=1):
        column_name = str(cell.value)
        if column_name in date_columns:
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row=row_index, column=column_index).number_format = "yyyy-mm-dd"

    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "A2"
    _finish_sheet(sheet)


def _format_dates(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    formatted = frame.copy()
    for column in ("latest_week_date", "current_signal_date"):
        if column in formatted.columns:
            formatted[column] = pd.to_datetime(formatted[column], errors="coerce").dt.date
    return formatted


def _excel_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    return value


def _finish_sheet(sheet: Any, widths: dict[str, int] | None = None) -> None:
    widths = widths or {}
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        if column_letter in widths:
            sheet.column_dimensions[column_letter].width = widths[column_letter]
            continue
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)
