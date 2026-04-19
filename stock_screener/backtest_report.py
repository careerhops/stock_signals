from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from stock_screener.backtest import BacktestResult


HEADER_FILL = PatternFill("solid", fgColor="DFF5EE")
TITLE_FILL = PatternFill("solid", fgColor="168A75")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)


def write_backtest_workbook(result: BacktestResult, workbook_path: Path) -> Path:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    _write_summary_sheet(summary_sheet, result.summary)
    _write_dataframe_sheet(workbook, "Stock Stats", result.stock_stats)
    _write_dataframe_sheet(workbook, "Closed Trades", _format_dates(result.trades))
    _write_dataframe_sheet(workbook, "Open Positions", _format_dates(result.open_positions))

    workbook.save(workbook_path)
    return workbook_path


def _write_summary_sheet(sheet: Any, summary: dict[str, Any]) -> None:
    sheet["A1"] = "BUY-to-next-SELL Backtest Summary"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = TITLE_FILL
    sheet.merge_cells("A1:D1")

    rows = [
        ("Exchange", summary.get("exchange", "")),
        ("Symbols processed", summary.get("symbols_processed", 0)),
        ("Symbols with closed trades", summary.get("symbols_with_closed_trades", 0)),
        ("Closed BUY-to-SELL trades", summary.get("closed_trades", 0)),
        ("Winning trades", summary.get("winning_trades", 0)),
        ("Losing trades", summary.get("losing_trades", 0)),
        ("Breakeven trades", summary.get("breakeven_trades", 0)),
        ("Open BUY positions", summary.get("open_positions", 0)),
        ("Win rate", summary.get("win_rate_pct", 0) / 100),
        ("Loss rate", summary.get("loss_rate_pct", 0) / 100),
        ("Average return", summary.get("avg_return_pct", 0) / 100),
        ("Median return", summary.get("median_return_pct", 0) / 100),
        ("Best return", summary.get("best_return_pct", 0) / 100),
        ("Worst return", summary.get("worst_return_pct", 0) / 100),
        ("Trades went up before SELL", summary.get("trades_went_up_before_sell", 0)),
        ("Went up before SELL rate", summary.get("went_up_before_sell_rate_pct", 0) / 100),
        ("Average max gain before SELL", summary.get("avg_max_gain_before_sell_pct", 0) / 100),
        ("Median max gain before SELL", summary.get("median_max_gain_before_sell_pct", 0) / 100),
        ("Hit 5% before SELL rate", summary.get("hit_5pct_before_sell_rate_pct", 0) / 100),
        ("Hit 10% before SELL rate", summary.get("hit_10pct_before_sell_rate_pct", 0) / 100),
        ("Hit 15% before SELL rate", summary.get("hit_15pct_before_sell_rate_pct", 0) / 100),
        ("Hit 20% before SELL rate", summary.get("hit_20pct_before_sell_rate_pct", 0) / 100),
    ]

    for row_index, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)
        sheet.cell(row=row_index, column=1).font = BOLD_FONT

    for row_index in range(11, 25):
        sheet.cell(row=row_index, column=2).number_format = "0.00%"

    chart_data_start = 7
    sheet["D6"] = "Outcome"
    sheet["E6"] = "Count"
    sheet["D7"] = "Wins"
    sheet["E7"] = summary.get("winning_trades", 0)
    sheet["D8"] = "Losses"
    sheet["E8"] = summary.get("losing_trades", 0)
    sheet["D9"] = "Breakeven"
    sheet["E9"] = summary.get("breakeven_trades", 0)
    for cell in sheet["D6:E6"][0]:
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL

    chart = BarChart()
    chart.title = "Closed Trade Outcomes"
    chart.y_axis.title = "Trades"
    chart.x_axis.title = "Outcome"
    data = Reference(sheet, min_col=5, min_row=6, max_row=9)
    cats = Reference(sheet, min_col=4, min_row=7, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    sheet.add_chart(chart, "D11")

    _finish_sheet(sheet, widths={"A": 30, "B": 18, "D": 18, "E": 12})


def _write_dataframe_sheet(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(title)
    if frame.empty:
        sheet.append(["No rows"])
        _finish_sheet(sheet)
        return

    sheet.append(list(frame.columns))
    for _, row in frame.iterrows():
        sheet.append([_excel_value(value) for value in row.tolist()])

    for cell in sheet[1]:
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    percent_columns = {
        "return_pct",
        "win_rate_pct",
        "loss_rate_pct",
        "avg_return_pct",
        "median_return_pct",
        "best_return_pct",
        "worst_return_pct",
        "open_return_pct",
        "max_gain_before_sell_pct",
        "went_up_before_sell_rate_pct",
        "avg_max_gain_before_sell_pct",
        "median_max_gain_before_sell_pct",
        "best_max_gain_before_sell_pct",
        "hit_5pct_before_sell_rate_pct",
        "hit_10pct_before_sell_rate_pct",
        "hit_15pct_before_sell_rate_pct",
        "hit_20pct_before_sell_rate_pct",
    }
    date_columns = {"buy_date", "sell_date", "latest_date", "max_gain_date"}
    for column_index, cell in enumerate(sheet[1], start=1):
        column_name = str(cell.value)
        if column_name in percent_columns:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=column_index).number_format = "0.00"
        if column_name in date_columns:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=column_index).number_format = "yyyy-mm-dd"

    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "A2"
    _finish_sheet(sheet)


def _finish_sheet(sheet: Any, widths: dict[str, int] | None = None) -> None:
    widths = widths or {}
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        if column_letter in widths:
            sheet.column_dimensions[column_letter].width = widths[column_letter]
            continue
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)


def _format_dates(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    formatted = frame.copy()
    for column in ("buy_date", "sell_date", "latest_date", "max_gain_date"):
        if column in formatted.columns:
            formatted[column] = pd.to_datetime(formatted[column], errors="coerce").dt.date
    return formatted


def _excel_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    return value
