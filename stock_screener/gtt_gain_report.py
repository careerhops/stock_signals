from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from stock_screener.gtt_gain_study import GttGainStudyResult


HEADER_FILL = PatternFill("solid", fgColor="DFF5EE")
TITLE_FILL = PatternFill("solid", fgColor="168A75")
SUBTITLE_FILL = PatternFill("solid", fgColor="E5F1FB")
BOLD_FONT = Font(bold=True)
WHITE_TITLE_FONT = Font(size=16, bold=True, color="FFFFFF")


def write_gtt_gain_workbook(result: GttGainStudyResult, workbook_path: Path) -> Path:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    _write_summary_sheet(summary_sheet, result.summary)
    _write_dataframe_sheet(workbook, "Stock GTT Stats", _format_dates(result.stock_stats))
    _write_dataframe_sheet(workbook, "Pair Details", _format_dates(result.pair_details))
    _write_dataframe_sheet(workbook, "Open BUY Positions", _format_dates(result.open_positions))

    workbook.save(workbook_path)
    return workbook_path


def _write_summary_sheet(sheet: Any, summary: dict[str, Any]) -> None:
    sheet["A1"] = "GTT Gain Study Summary"
    sheet["A1"].font = WHITE_TITLE_FONT
    sheet["A1"].fill = TITLE_FILL
    sheet.merge_cells("A1:D1")

    rows = [
        ("Exchange", summary.get("exchange", "")),
        ("NSE symbols processed", summary.get("symbols_processed", 0)),
        ("Closed BUY-to-SELL pairs", summary.get("closed_pairs", 0)),
        ("Valid pairs with daily data", summary.get("valid_pairs", 0)),
        ("Pairs without daily data", summary.get("pairs_without_daily_window", 0)),
        ("Open BUY positions", summary.get("open_buy_positions", 0)),
        ("Overall median max gain", summary.get("overall_median_max_gain_pct", 0) / 100),
        ("Overall average max gain", summary.get("overall_avg_max_gain_pct", 0) / 100),
        ("Pairs went above BUY price", summary.get("pairs_went_above_buy_price", 0)),
        ("Went above BUY price rate", summary.get("went_above_buy_price_rate_pct", 0) / 100),
        ("Hit 5% rate", summary.get("hit_5pct_rate_pct", 0) / 100),
        ("Hit 10% rate", summary.get("hit_10pct_rate_pct", 0) / 100),
        ("Hit 15% rate", summary.get("hit_15pct_rate_pct", 0) / 100),
        ("Hit 20% rate", summary.get("hit_20pct_rate_pct", 0) / 100),
        ("Hit 25% rate", summary.get("hit_25pct_rate_pct", 0) / 100),
        ("Hit 30% rate", summary.get("hit_30pct_rate_pct", 0) / 100),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)
        sheet.cell(row=row_index, column=1).font = BOLD_FONT

    for row_index in range(9, 19):
        sheet.cell(row=row_index, column=2).number_format = "0.00%"

    sheet["D4"] = "Target"
    sheet["E4"] = "Hit rate"
    for threshold, row_index in zip((5, 10, 15, 20, 25, 30), range(5, 11)):
        sheet.cell(row=row_index, column=4, value=f"{threshold}%")
        sheet.cell(row=row_index, column=5, value=summary.get(f"hit_{threshold}pct_rate_pct", 0) / 100)
        sheet.cell(row=row_index, column=5).number_format = "0.00%"

    for cell in sheet["D4:E4"][0]:
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL

    chart = BarChart()
    chart.title = "Historical GTT Target Hit Rates"
    chart.y_axis.title = "Hit Rate"
    chart.x_axis.title = "Target"
    data = Reference(sheet, min_col=5, min_row=4, max_row=10)
    cats = Reference(sheet, min_col=4, min_row=5, max_row=10)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    sheet.add_chart(chart, "D12")

    _finish_sheet(sheet, widths={"A": 32, "B": 18, "D": 14, "E": 14})


def _write_dataframe_sheet(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(title)
    if frame.empty:
        sheet.append(["No rows"])
        _finish_sheet(sheet)
        return

    sheet.append(list(frame.columns))
    for _, row in frame.iterrows():
        sheet.append([_excel_value(value) for value in row.tolist()])

    for cell in sheet[1]:
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    percent_columns = {
        "buy_to_sell_return_pct",
        "max_gain_pct",
        "went_above_buy_price_rate_pct",
        "median_max_gain_pct",
        "avg_max_gain_pct",
        "best_max_gain_pct",
        "hit_5pct_rate_pct",
        "hit_10pct_rate_pct",
        "hit_15pct_rate_pct",
        "hit_20pct_rate_pct",
        "hit_25pct_rate_pct",
        "hit_30pct_rate_pct",
        "suggested_conservative_gtt_pct",
        "suggested_moderate_gtt_pct",
        "open_max_gain_pct",
    }
    date_columns = {"buy_date", "sell_date", "highest_price_date", "latest_date", "latest_week_date", "latest_signal_date"}
    for column_index, cell in enumerate(sheet[1], start=1):
        column_name = str(cell.value)
        if column_name in percent_columns:
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row=row_index, column=column_index).number_format = "0.00"
        if column_name in date_columns:
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row=row_index, column=column_index).number_format = "yyyy-mm-dd"

    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "A2"
    _finish_sheet(sheet)


def _format_dates(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    formatted = frame.copy()
    for column in ("buy_date", "sell_date", "highest_price_date", "latest_date"):
        if column in formatted.columns:
            formatted[column] = pd.to_datetime(formatted[column], errors="coerce").dt.date
    return formatted


def _excel_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    return value


def _finish_sheet(sheet: Any, widths: dict[str, int] | None = None) -> None:
    widths = widths or {}
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        if column_letter in widths:
            sheet.column_dimensions[column_letter].width = widths[column_letter]
            continue
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)
